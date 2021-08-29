from Produto import Produto
from openpyxl import load_workbook

class Excel:
    def __init__(self, caminhoDoArquivo):
        self._excelFile = load_workbook(caminhoDoArquivo)
        self._caminhoDoArquivo = caminhoDoArquivo
        self._planilhaAtiva = self._excelFile.active
    
    def retornaTodaTabela(self):

        #Seta a quantidade de linhas e colunas da tabela
        linhas = self._planilhaAtiva.max_row
        colunas = self._planilhaAtiva.max_column

        #Inicializa variaveis das linhas e do conteúdo da tabela.
        valoresDasLinhas = []
        conteudoDaPlanilha = []

        #Busca conteúdo de toda a tabela
        for i in range(1, linhas + 1):
            valoresDasLinhas = []
            for j in range(1, colunas + 1):
                celula = self._planilhaAtiva.cell(row=i, column=j)
                valoresDasLinhas.append(celula.value)

            codigo      = valoresDasLinhas[0]
            descricao   = valoresDasLinhas[1]
            precoCompra = valoresDasLinhas[2]
            precoVenda  = valoresDasLinhas[3]
            quantidade  = valoresDasLinhas[4]

            produto = Produto(codigo, descricao, precoCompra, precoVenda, quantidade)
            conteudoDaPlanilha.append(produto)
        return conteudoDaPlanilha
    
    def buscaProduto(self, produto):
        tabela = self.retornaTodaTabela()
        produtoEncontrado = ""
        for i in range(len(tabela)):
            if(produto == tabela[i].codigo):
                produtoEncontrado = tabela[i]
        return produtoEncontrado

    def adicionaProduto(self, produto):
        #Seta a quantidade de linhas e colunas da tabela
        CAMINHO_DO_ARQUIVO = 'inventario-loja.xlsx'
        ultimaLinha     = self._planilhaAtiva.max_row
        ultimaColuna    = self._planilhaAtiva.max_column
        for j in range(1, ultimaColuna + 1):
            celula = self._planilhaAtiva.cell(row=ultimaLinha+1, column=j)
            celula.value = "TESTE1"
        self._excelFile.save(self._caminhoDoArquivo)
    
    def atualizaQuantidadeProduto(self, produto, quantidade: int):
        quantidadeDeLinhas = self._planilhaAtiva.max_row
        colunaDaQuantidade = 5
        colunaComCodigoDoProduto = 1

        indiceDaLinhaDoPruduto = 0
        #Busca linha do produto
        for i in range(1, quantidadeDeLinhas + 1):
            celula = self._planilhaAtiva.cell(row=i, column=colunaComCodigoDoProduto)
            if(celula.value == produto):
                indiceDaLinhaDoPruduto = i

        #Altera linha do produto
        celula = self._planilhaAtiva.cell(row=indiceDaLinhaDoPruduto, column=colunaDaQuantidade)
        valorCelula = int(celula.value)
        if(valorCelula - quantidade < 0):
            return False
        else:
            celula.value = valorCelula - quantidade
            self._excelFile.save(self._caminhoDoArquivo)
            return True


